import logging
import math
import warnings
import pandas as pd
import time
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from copy import copy
from datetime import datetime, timedelta
from openpyxl.styles import Alignment
from setting_wiz import *
import setting_wiz


def fillter_rec():
    finder_file = find_file(Data_path)
    file_data = finder_file.file_last_time()
    # wb = load_workbook(file_data)
    # ws = wb.active
    df = pd.read_excel(file_data)

    df['Call Time'] = pd.to_datetime(df['Call Time'], dayfirst= True)
    
    duplicate_numbers = df[df.duplicated(subset=['Contact Number'], keep=False)]
    if not duplicate_numbers.empty:
        print("Duplicate Contact Numbers:")
        print(duplicate_numbers[['Contact Number', 'Call Time']])

    df_latest = df.sort_values('Call Time', ascending=False).drop_duplicates(subset=['Contact Number'], keep='first')
    df_latest['Call Time'] = df_latest['Call Time'].dt.strftime('%d/%m/%Y %H:%M:%S')
    
    
    # for row_idx, row_data in enumerate(df_latest.values, start=2): 
    #     for col_idx, value in enumerate(row_data, start=1):
    #         ws.cell(row=row_idx, column=col_idx, value=value)
    #         print('write success')
    
    
    new_filename = os.path.join(sub1_folder, os.path.basename(file_data).replace(".xlsx", f"_Fill_Rec.xlsx"))
    df_latest.to_excel(new_filename, index= False)
    # wb.save(new_filename)
    
    
def cut_data():
    finder_file = find_file(Data_path)
    file_data = finder_file.file_last_time()
    wb = load_workbook(file_data)
    ws = wb.active
    df = pd.read_excel(file_data)

def Create_folder_BC():
    finder_file = find_file(Fill_rec_path)
    file_data = finder_file.file_last_time()
    df = pd.read_excel(file_data)

    current_date = date.strftime("%d-%m-%Y") # เชื่อมวันรันตรงนี้

    date_folder_path = os.path.join(sub2_folder, current_date)

    if not os.path.exists(date_folder_path):
        os.makedirs(date_folder_path)
        print(f'Create date folder: {date_folder_path}.')
    else:
        print(f'{date_folder_path} already exists.')

    
    unique_task_name = df['Task Name'].unique()

    for task_name in unique_task_name:

        parts =  task_name.split('_')
        main_fd = parts[0]
        sub_fd = parts[1]

        main_folder_path = os.path.join(date_folder_path, main_fd)
        if not os.path.exists(main_folder_path):
            os.makedirs(main_folder_path)
            print(f'create main folder {main_folder_path}.')
        else:
            print(f'{main_folder_path} is already exists.')
        
        sub_folder_path = os.path.join(main_folder_path,sub_fd)

        if not os.path.exists(sub_folder_path):
            os.makedirs(sub_folder_path)
            print(f'Create Folder: {sub_folder_path}.')
        else:
            print(f'{sub_folder_path} folder already exists.')
        Put_data(df, task_name, sub_folder_path)

def Put_data(dataframe, task_name,sub_f_p ):
    task_data = dataframe[dataframe['Task Name'] == task_name]

    file_name = f"{task_name}.xlsx"
    file_path = os.path.join(sub_f_p, file_name)
    
    task_data.to_excel(file_path, index= False)
    print(f"Saved data for {task_name} in {file_path}.")

def map_data():
    fillter_rec()
    Create_folder_BC()